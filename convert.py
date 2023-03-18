from msilib.schema import Error
import os
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox
from tkinter import font

import pdfplumber
import pandas as pd
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.protection import SheetProtection

#definition to move worksheet by index, not offset:(
def movedirect(workbook,worksheet,idx):
    #find worksheet index
    orgloc = workbook.index(worksheet)
    delta = idx - orgloc
    workbook.move_sheet(worksheet,offset = delta)

class CONVERT:
    
    def __init__(self,master):
        self.master = master
        self.master.title("土城醫院檢驗科")  
        self.master.geometry('600x400')
        self.master.config(background='#323232')
        self.master.iconbitmap("blood.ico")
        style = ttk.Style()
        style.configure("title.TLabel",font=('微軟正黑體',26), foreground="#FFFFFF", background="#323232")
        style.configure("path.TLabel", font=('微軟正黑體',16),foreground="#FFFFFF", background="#323232")
        style.configure("head2.TLabel", font=('微軟正黑體',18),foreground="#FFFFFF", background="#323232")
        style.configure("cc.TLabel", font=('微軟正黑體',8),foreground="#FFFFFF", background="#323232")
        style.configure("browser.TButton", font=('微軟正黑體',10),foreground="#000000", background="#323232")
        style.configure("yn.TButton", font=('微軟正黑體',16),foreground="#000000", background="#323232")
        self.master.bind('<Return>', self.callback)
        #建立標題
        self.label_title = ttk.Label(self.master, text='新批號輸入',style="title.TLabel")  
        self.pdf_label = ttk.Label(self.master, text='輸入 => Panel_A PDF ',style="head2.TLabel")  
        #建立一個`label`名為`label_path1: `  
        self.label_path1 = ttk.Label(self.master, text='路徑： ',style="path.TLabel")  
        self.targetexcel_label = ttk.Label(self.master, text='輸入 => 不規則抗體excel檔 ',style="head2.TLabel")  
        self.outputexcel_label = ttk.Label(self.master, text='輸出 => 存檔位置 ',style="head2.TLabel")  
        #建立一個`label`名為`label_path2: `  
        self.label_path2 = ttk.Label(self.master, text='路徑： ',style="path.TLabel")
        #建立一個`label`名為`label_path3: `  
        self.label_path3 = ttk.Label(self.master, text='路徑： ',style="path.TLabel")
        # 建立一個path1輸入框,並設定尺寸  
        self.input_path1 = ttk.Entry(self.master,width=35)
        # 建立一個path2輸入框,並設定尺寸  
        self.input_path2 = ttk.Entry(self.master,width=35)  
        # 建立一個path3輸入框,並設定尺寸  
        self.input_path3 = ttk.Entry(self.master,width=35)  
        #瀏覽按鈕
        self.browse1_button = ttk.Button(self.master, command = self.browse_pdf, text = "瀏覽...", width=4,style="browser.TButton")
        self.browse2_button = ttk.Button(self.master, command = self.browse_xls, text = "瀏覽...", width=4,style="browser.TButton")
        self.browse3_button = ttk.Button(self.master, command = self.output_xls, text = "瀏覽...", width=4,style="browser.TButton")
        # 建立一個登入系統的按鈕  
        self.login_button = ttk.Button(self.master, command = self.OK, text = "確定", width=6,style="yn.TButton")
        self.login_button.pack()
        # 建立一個退出系統的按鈕  
        self.exit_button = ttk.Button(self.master, command = self.exit, text = "退出", width=6,style="yn.TButton")
        self.exit_button.pack()
        self.cc = ttk.Label(
            self.master, 
            text='@Design by Henry Tsai',
            style="cc.TLabel",
            )  
        self.label_title.place(relx=0.5, rely=0.1, anchor=tk.CENTER) 
        self.pdf_label.place(relx=0.05, rely=0.2, anchor='w')
        self.targetexcel_label.place(relx=0.05, rely=0.42, anchor='w')
        self.outputexcel_label.place(relx=0.05, rely=0.64, anchor='w')
        self.label_path1.place(relx=0.22, rely=0.31, anchor=tk.CENTER)  
        self.label_path2.place(relx=0.22, rely=0.53, anchor=tk.CENTER)  
        self.label_path3.place(relx=0.22, rely=0.75, anchor=tk.CENTER)
        self.input_path1.place(relx=0.49, rely=0.31, anchor=tk.CENTER)  
        self.input_path2.place(relx=0.49, rely=0.53, anchor=tk.CENTER)
        self.input_path3.place(relx=0.49, rely=0.75, anchor=tk.CENTER)
        self.browse1_button.place(relx=0.75, rely=0.31, anchor=tk.CENTER)
        self.browse2_button.place(relx=0.75, rely=0.53, anchor=tk.CENTER)
        self.browse3_button.place(relx=0.75, rely=0.75, anchor=tk.CENTER)
        self.login_button.place(relx=0.6, rely=0.9, anchor=tk.CENTER)  
        self.exit_button.place(relx=0.4, rely=0.9, anchor=tk.CENTER)
        self.cc.place(relx=1, rely=1,anchor=tk.SE)
    def browse_pdf(self):
        self.filename  = filedialog.askopenfilename(initialdir="E:",title="選擇檔案",filetypes=(("pdf","*.pdf"),("all files","*.*")))
        self.input_path1.delete(0,END)
        self.input_path1.insert(0,self.filename)
    def browse_xls(self):
        self.filename2  = filedialog.askopenfilename(initialdir="E:",title="選擇檔案",filetypes=(("Excel VBA","*.xlsm"),("Excel","*.xlsx"),("CSV UTF-8","*.csv"),("Excel 2003","*.xls"),("all files","*.*")))
        self.input_path2.delete(0,END)
        self.input_path2.insert(0,self.filename2)
    def output_xls(self):
        self.folder  = filedialog.askdirectory(initialdir="E:",title="選擇存檔路徑")
        self.input_path3.delete(0,END)
        self.input_path3.insert(0,self.folder)
    def exit(self):
        self.master.destroy()
    def callback(self, event):
        self.OK()
    def OK(self):
        # doc = "./RA209.pdf"
        path1 = self.input_path1.get()
        path2 = self.input_path2.get()
        path3 = self.input_path3.get()
        ##這裡要加入驗證動作
        try:
            pdf = pdfplumber.open(path1)
        except FileNotFoundError:
            tk.messagebox.showerror(title='長庚土城檢驗科', message='Panel_A pdf 檔案路徑讀取錯誤，請檢查!')
            return
        # page0 = pdf.pages[0]
        rawdata=[]
        rawdata_text=[]
        # table = page0.extract_table()
        # pagetext = str(page0.extract_text()).split(" ")
        # lotno = pagetext[4]
        try:
            page0 = pdf.pages[0]
            table = page0.extract_table()
            pagetext = str(page0.extract_text()).split(" ")
            lotno = pagetext[4]
        except:    
            tk.messagebox.showerror(title='長庚土城檢驗科', message='pdf 並非Panel_A檔案，請檢查!')
            return
        if lotno[0:2] != "RA":
            tk.messagebox.showerror(title='長庚土城檢驗科', message='pdf 並非Panel_A檔案，請檢查!')
            return
        # print(lotno)
        #1-32 -> 0,32
        for i in range(1,13):
            del table[i][29]
            rawdata.append(table[i][0:31])
        df = pd.DataFrame(rawdata)
        #row1 to header
        df.columns = df.iloc[0]
        df = df[1:]
        # print(df)
        #refresh data ["/"" -> "?""],["+s" -> "+"]
        df["Jsa"] = df["Jsa"].replace("/","?")
        df["P1"] = df["P1"].replace("+s","+")
        df = df.replace("0","-")
        # print(df)
        ##edit xlsm
        # wb = load_workbook(filename='./不規則抗體.xlsm', read_only=False, keep_vba=True)
        try:
            wb = load_workbook(filename=path2, read_only=False, keep_vba=True)
            temp = wb["temp"]
        except openpyxl.utils.exceptions.InvalidFileException:
            tk.messagebox.showerror(title='長庚土城檢驗科', message='不規則抗體excel檔案路徑讀取錯誤，請檢查!')
            return
        except KeyError:
            tk.messagebox.showerror(title='長庚土城檢驗科', message='xls 並非不規則抗體excel檔案，請檢查!')
            return
        #copy temp
        ws = wb.copy_worksheet(temp)
        ws.title = "Lot. %s"%(lotno)
        #copy conditional_formatting
        ws.conditional_formatting._cf_rules = temp.conditional_formatting._cf_rules.copy()
        #move copy lot
        movedirect(wb,ws,2)
        ##edit value
        #paste lot No
        for i in range(12,23):
            ws.cell(column=1,row=i).value = lotno
        #paste +/- ["D12 : AE22"] => [(4,12):(31,22)]
        rawdata = df.drop(columns=["Cell#","Rh-hr","Donor\nNumber"])
        rawdata = rawdata.to_numpy().tolist()
        for row in range(12,23):
            for col in range(4,32):
                ws.cell(column = col,row = row).value = rawdata[row - 12][col - 4]
        #protect ws
        protection = SheetProtection(
            sheet=True,
            password='1234',
            sort=True,
            insertRows=True,
            insertColumns=True,
            deleteRows=True,
            deleteColumns=True,
            selectLockedCells=True,
            selectUnlockedCells=False,
        )
        ws.protection = protection 
        ##檢查有沒有重複檔名的檔案，askyesno要不要取代
        filepath = '%s/不規則抗體.xlsm'%(path3)
        if os.path.isfile(filepath):
            if tk.messagebox.askyesno(title='長庚土城檢驗科', message='輸出資料夾已有相同檔名的檔案，是否取代?'):
                wb.save(filepath)
            else:
                filepath = '%s/不規則抗體_new.xlsm'%(path3)
                wb.save(filepath)
                tk.messagebox.showinfo(title='長庚土城檢驗科', message='檔案另存成功!')
                return
        else:
            wb.save(filepath)
        ##檢查是否輸出成功，msgbox.showinfo
        if os.path.isfile(filepath):
            tk.messagebox.showinfo(title='長庚土城檢驗科', message='檔案新增成功!')
            self.input_path1.delete(0,END)
            self.input_path2.delete(0,END)
            self.input_path3.delete(0,END)
        else:
            tk.messagebox.showinfo(title='長庚土城檢驗科', message='檔案新增失敗QQ')
def main():  
    # 初始化物件  
    root =Tk()
    L = CONVERT(root)  
    root.mainloop()

if __name__ == '__main__':  
    main()